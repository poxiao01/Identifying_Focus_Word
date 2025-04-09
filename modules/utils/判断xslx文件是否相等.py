import openpyxl

def are_excel_files_equal(file1, file2):
    """
    比较两个 Excel 文件是否完全一致，并输出具体不一致的行和列信息。

    参数：
        file1 -- 第一个 Excel 文件路径
        file2 -- 第二个 Excel 文件路径

    返回：
        tuple: (是否一致的布尔值, 不一致详情列表)
    """
    differences = []
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # 比较工作表数量
    if len(wb1.sheetnames) != len(wb2.sheetnames):
        differences.append(f"工作表数量不同：file1 有 {len(wb1.sheetnames)} 个, file2 有 {len(wb2.sheetnames)} 个。")

    # 遍历 file1 中的所有工作表
    for sheet_name in wb1.sheetnames:
        if sheet_name not in wb2.sheetnames:
            differences.append(f"工作表 '{sheet_name}' 存在于 file1 中，但不存在于 file2 中。")
            continue

        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # 取两个工作表中较大的行数和列数
        max_row = max(sheet1.max_row, sheet2.max_row)
        max_col = max(sheet1.max_column, sheet2.max_column)

        for row in range(1, max_row + 1):
            row_diff = []
            for col in range(1, max_col + 1):
                val1 = sheet1.cell(row=row, column=col).value
                val2 = sheet2.cell(row=row, column=col).value
                if val1 != val2:
                    row_diff.append(f"列 {col}（file1='{val1}' vs file2='{val2}'）")
            if row_diff:
                differences.append(f"工作表 '{sheet_name}' 第 {row} 行不一致： " + "; ".join(row_diff))

    if differences:
        return False, differences
    else:
        return True, ["两个文件完全一致"]

# 使用示例
file1 = "F:/ExperimentalC_code/BetterCode/data/results/experiment_results.xlsx"
file2 = "F:/ExperimentalC_code/BetterCode/data/results/前缀树_experiment_results.xlsx"

equal, details = are_excel_files_equal(file1, file2)
print(f"两个文件是否完全相等: {equal}")
if not equal:
    print("具体不一致的行如下：")
    for detail in details:
         print(detail)
else:
    print("两个文件完全一致！")
